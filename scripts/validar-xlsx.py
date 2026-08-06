"""Abre los `.xlsx` que exportamos y dice si Excel podría abrirlos SIN QUEJARSE.

Lo invoca `scripts/validar-xlsx.mjs`. Es el gemelo de `validar-dxf.py` y de
`validar-xsd.py`, y existe por la misma razón que los dos: **un oráculo que no es
nuestro**.

── ⛔ POR QUÉ HAY DOS PASADAS, Y LA SEGUNDA NO TOCA openpyxl ────────────────
Porque la lección más cara de este proyecto es exactamente esa. El **2026-08-05**
un usuario abrió en ZWCAD 2023 el DXF que la aplicación exportaba y el programa se
quedó en blanco y bloqueado; `ezdxf` daba verde porque **rellena por su cuenta lo
que falta al cargar**, así que preguntarle si el fichero estaba completo respondía
por su modelo y no por el fichero.

**openpyxl hace lo mismo.** Si falta una relación, si un índice de estilo apunta
fuera de la tabla o si una parte no está declarada en `[Content_Types].xml`,
openpyxl la mayoría de las veces lo ignora o se lo inventa — y Excel, en cambio,
enseña «Hemos encontrado un problema con el contenido de este archivo». Así que:

  · **Pasada 1 (openpyxl)**: ¿se puede LEER? ¿los números son números, los
    formatos están, las pestañas se llaman como tienen que llamarse?
  · **Pasada 2 (solo la biblioteca estándar)**: ¿el paquete está BIEN FORMADO?
    CRC de cada entrada, partes declaradas y presentes, relaciones que apuntan a
    algo, índices de estilo dentro de rango.

Ninguna sustituye a la otra. **Y ninguna sustituye a abrir el fichero en Excel**:
lo del DXF lo destapó una persona, no una máquina.

── LO QUE COMPRUEBA LA PASADA 2, Y POR QUÉ CADA COSA ───────────────────────
  1. **El CRC de cada entrada** (`testzip`). Un ZIP con un CRC mal es un fichero
     que Excel rechaza entero, y es justo lo que pasaría si los bytes se hubieran
     recodificado por el camino.
  2. **Todas las entradas sin comprimir.** No es un requisito del formato: es la
     decisión de F20 (`export/xlsx.js`), y si alguien la cambia sin querer, aquí
     se entera.
  3. ⭐ **Cada parte del ZIP está declarada en `[Content_Types].xml`**, por su
     extensión (`Default`) o por su nombre (`Override`); y cada `Override`
     declarado existe de verdad en el ZIP. Es el error de OOXML que openpyxl
     perdona y Excel no.
  4. ⭐ **Cada `r:id` que nombra `workbook.xml` existe en `workbook.xml.rels`**, y
     apunta a una parte que está. Si esto se descoloca, Excel abre la hoja
     equivocada bajo cada pestaña: el fichero no se rompe, solo se vuelve mentira.
  5. ⭐ **Cada `s="N"` de una hoja cae dentro de `cellXfs`.** Un índice de estilo
     fuera de rango es «archivo dañado» en Excel y un encogimiento de hombros en
     openpyxl.
  6. **No hay comentario de ZIP** y el directorio central acaba donde empieza el
     EOCD.

Uso:  python scripts/validar-xlsx.py <manifiesto.json>

Códigos de salida:
  0 → todo bien, y las DOS pasadas han corrido
  1 → algo está mal (arriba se dice qué)
  2 → no se ha podido ni empezar
  3 → la pasada 2 ha ido bien pero NO hay openpyxl: media medición
"""

import json
import re
import sys
import zipfile

# ⚠️ La consola de Windows no viene en UTF-8, así que sin esto los mensajes de este
# script salen con la codificación cambiada («Pasada 2 � paquete») — medido al
# escribirlo. Un validador cuyo informe no se puede leer es medio validador.
if hasattr(sys.stdout, "reconfigure"):
    sys.stdout.reconfigure(encoding="utf-8", errors="replace")

NS_REL = "{http://schemas.openxmlformats.org/package/2006/relationships}"
NS_HOJA = "{http://schemas.openxmlformats.org/spreadsheetml/2006/main}"
NS_TIPOS = "{http://schemas.openxmlformats.org/package/2006/content-types}"

try:
    import xml.etree.ElementTree as ET
except ImportError:  # pragma: no cover — no pasa en un Python de verdad
    print("  No hay xml.etree en este Python.")
    sys.exit(2)


class Problemas:
    """Acumula lo que va mal, con el fichero al que le pasa."""

    def __init__(self):
        self.lista = []

    def add(self, fichero, texto):
        self.lista.append((fichero, texto))

    def __len__(self):
        return len(self.lista)


# ═══════════════════════════════════════════════════════════════════════════
# PASADA 2 · el paquete, con la biblioteca estándar y NADA MÁS
# ═══════════════════════════════════════════════════════════════════════════


def pasada_estructural(ruta, esperado, problemas):
    nombre = esperado["nombre"]

    with open(ruta, "rb") as f:
        crudo = f.read()

    if not crudo.startswith(b"PK\x03\x04"):
        problemas.add(nombre, "no empieza por la firma de un ZIP (PK\\x03\\x04)")
        return

    # El EOCD sin comentario son los últimos 22 bytes.
    if crudo[-22:-18] != b"PK\x05\x06":
        problemas.add(nombre, "no acaba en un EOCD sin comentario: hay cola de más")
        return

    try:
        z = zipfile.ZipFile(ruta)
    except zipfile.BadZipFile as e:
        problemas.add(nombre, f"zipfile no lo abre: {e}")
        return

    # 1 · CRC de cada entrada. `testzip` devuelve el nombre del PRIMERO que falla.
    malo = z.testzip()
    if malo is not None:
        problemas.add(nombre, f"el CRC de «{malo}» no cuadra con sus bytes")

    partes = {i.filename for i in z.infolist()}

    # 2 · Todas sin comprimir: la decisión de F20.
    comprimidas = [i.filename for i in z.infolist() if i.compress_type != 0]
    if comprimidas:
        problemas.add(nombre, f"hay entradas comprimidas y F20 las emite en STORE: {comprimidas}")

    # 3 · Content_Types: lo declarado existe, y lo que existe está declarado.
    if "[Content_Types].xml" not in partes:
        problemas.add(nombre, "no trae [Content_Types].xml, que es obligatorio en OOXML")
        return

    tipos = ET.fromstring(z.read("[Content_Types].xml"))
    defaults = {d.get("Extension", "").lower() for d in tipos.findall(f"{NS_TIPOS}Default")}
    overrides = {o.get("PartName", "").lstrip("/") for o in tipos.findall(f"{NS_TIPOS}Override")}

    for declarada in overrides:
        if declarada not in partes:
            problemas.add(nombre, f"[Content_Types].xml declara «{declarada}» y no está en el ZIP")

    for parte in partes:
        if parte == "[Content_Types].xml" or parte in overrides:
            continue
        extension = parte.rsplit(".", 1)[-1].lower() if "." in parte else ""
        if extension not in defaults:
            problemas.add(
                nombre,
                f"«{parte}» no está declarada: ni por Override ni por la extensión «{extension}»",
            )

    # 4 · Las relaciones del libro apuntan a algo que existe.
    if "xl/workbook.xml" not in partes or "xl/_rels/workbook.xml.rels" not in partes:
        problemas.add(nombre, "falta xl/workbook.xml o su fichero de relaciones")
        return

    rels = ET.fromstring(z.read("xl/_rels/workbook.xml.rels"))
    destino_de = {}
    for r in rels.findall(f"{NS_REL}Relationship"):
        destino = r.get("Target", "")
        completa = destino if destino.startswith("xl/") else f"xl/{destino}"
        destino_de[r.get("Id")] = completa
        if completa not in partes:
            problemas.add(nombre, f"la relación {r.get('Id')} apunta a «{completa}», que no está")

    libro = ET.fromstring(z.read("xl/workbook.xml"))
    hojas_declaradas = []
    for hoja in libro.iter(f"{NS_HOJA}sheet"):
        rid = hoja.get(f"{{http://schemas.openxmlformats.org/officeDocument/2006/relationships}}id")
        if rid not in destino_de:
            problemas.add(nombre, f"la pestaña «{hoja.get('name')}» usa {rid}, que no está declarado")
            continue
        hojas_declaradas.append((hoja.get("name"), destino_de[rid]))

    nombres = [n for n, _ in hojas_declaradas]
    if nombres != esperado["hojas"]:
        problemas.add(nombre, f"las pestañas son {nombres} y se esperaban {esperado['hojas']}")

    # 5 · Los índices de estilo caen dentro de cellXfs.
    if "xl/styles.xml" not in partes:
        problemas.add(nombre, "falta xl/styles.xml")
        return
    estilos = z.read("xl/styles.xml").decode("utf-8")
    m = re.search(r'<cellXfs count="(\d+)"', estilos)
    if m is None:
        problemas.add(nombre, "styles.xml no declara cellXfs")
        return
    cuantos = int(m.group(1))
    reales = estilos.count("<xf ", estilos.index("<cellXfs"))
    if reales != cuantos:
        problemas.add(nombre, f"cellXfs dice {cuantos} y hay {reales} de verdad")

    for _, parte in hojas_declaradas:
        xml = z.read(parte).decode("utf-8")
        for usado in {int(s) for s in re.findall(r'<c [^>]*s="(\d+)"', xml)}:
            if usado >= cuantos:
                problemas.add(
                    nombre, f"«{parte}» usa el estilo {usado} y cellXfs solo declara {cuantos}"
                )


# ═══════════════════════════════════════════════════════════════════════════
# PASADA 1 · openpyxl: ¿se puede LEER, y dice lo que tiene que decir?
# ═══════════════════════════════════════════════════════════════════════════


def pasada_openpyxl(ruta, esperado, problemas):
    import openpyxl

    nombre = esperado["nombre"]
    try:
        wb = openpyxl.load_workbook(ruta)
    except Exception as e:  # noqa: BLE001 — cualquier cosa aquí es un fallo del fichero
        problemas.add(nombre, f"openpyxl no lo abre: {type(e).__name__}: {e}")
        return

    if wb.sheetnames != esperado["hojas"]:
        problemas.add(nombre, f"openpyxl ve las pestañas {wb.sheetnames}, no {esperado['hojas']}")
        return

    ws = wb[esperado["hojas"][0]]

    # El título de la caja y su combinación.
    if ws["A1"].value != "Coordenadas Parcela":
        problemas.add(nombre, f"A1 dice {ws['A1'].value!r} y no «Coordenadas Parcela»")
    if "A1:C1" not in [str(r) for r in ws.merged_cells.ranges]:
        problemas.add(nombre, "el título no está combinado en A1:C1")

    # ⭐ Lo que es media fase: las coordenadas tienen que ser NÚMEROS.
    primer = esperado.get("primerVertice")
    if primer is not None:
        for ref, valor in (("A", primer["n"]), ("B", primer["x"]), ("C", primer["y"])):
            celda = ws[f"{ref}{primer['fila']}"]
            if not isinstance(celda.value, (int, float)):
                problemas.add(
                    nombre,
                    f"{ref}{primer['fila']} vale {celda.value!r} ({type(celda.value).__name__}) "
                    "y tiene que ser un número: una coordenada en texto no se puede sumar",
                )
            elif abs(float(celda.value) - valor) > 1e-9:
                problemas.add(nombre, f"{ref}{primer['fila']} vale {celda.value} y se esperaba {valor}")

        fmt = ws[f"B{primer['fila']}"].number_format
        if fmt != "0.00":
            problemas.add(nombre, f"la coordenada lleva el formato {fmt!r} y no '0.00'")

    # Las unidades van en el FORMATO, nunca dentro del valor.
    for fila in ws.iter_rows():
        for celda in fila:
            if isinstance(celda.value, str) and celda.value.endswith(" m²") and celda.row > 1:
                problemas.add(
                    nombre,
                    f"{celda.coordinate} lleva «m²» DENTRO del texto: esa celda ya no se puede sumar",
                )

    # Y el libro no dictamina (regla de oro 9).
    prohibidas = ("válid", "inválid", "correct", "incorrect", "cumple", "conforme")
    for hoja in wb.worksheets:
        for fila in hoja.iter_rows():
            for celda in fila:
                if isinstance(celda.value, str):
                    bajo = celda.value.lower()
                    for p in prohibidas:
                        if p in bajo:
                            problemas.add(
                                nombre, f"{hoja.title}!{celda.coordinate} emite un veredicto: «{p}»"
                            )


# ═══════════════════════════════════════════════════════════════════════════


def main(argv):
    if len(argv) != 1:
        print("  Uso: python scripts/validar-xlsx.py <manifiesto.json>")
        return 2

    with open(argv[0], encoding="utf-8") as f:
        manifiesto = json.load(f)

    try:
        import openpyxl  # noqa: F401

        hay_openpyxl = True
    except ImportError:
        hay_openpyxl = False

    problemas = Problemas()

    for caso in manifiesto:
        pasada_estructural(caso["ruta"], caso, problemas)
        if hay_openpyxl:
            pasada_openpyxl(caso["ruta"], caso, problemas)

    print(f"  Pasada 2 · paquete      {len(manifiesto)} ficheros, con la biblioteca estándar")
    if hay_openpyxl:
        import openpyxl

        print(f"  Pasada 1 · lectura      openpyxl {openpyxl.__version__}")
    else:
        print("  Pasada 1 · lectura      ⚠️ SALTADA: no hay openpyxl en este Python")
        print("                          se instala con:  python -m pip install openpyxl")

    if len(problemas) > 0:
        print()
        for fichero, texto in problemas.lista:
            print(f"  ⛔ {fichero}: {texto}")
        return 1

    return 0 if hay_openpyxl else 3


if __name__ == "__main__":
    sys.exit(main(sys.argv[1:]))
